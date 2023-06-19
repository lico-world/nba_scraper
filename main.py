from time import sleep
from urllib.request import urlopen
from bs4 import BeautifulSoup
from openpyxl import Workbook
from datetime import datetime
import os

import private_constants
import config

# Path were you want to export data
PATH = private_constants.path

def get_team_acronym(team, year):
    critical_acronyms = ['NOP', 'NOH', 'NOK', 'NJN', 'BRK', 'CHO', 'CHA', 'CHH', 'SEA', 'OKC']
    if team not in critical_acronyms:
        return team
    else:
        if team == 'NOP' or team == 'NOH' or team == 'NOK':
            if year < 2006 or 2007 < year < 2014:
                return 'NOH'
            elif 2005 < year < 2008:
                return 'NOK'
            else:
                return 'NOP'

        elif team == 'NJN' or team == 'BRK':
            if year < 2013:
                return 'NJN'
            else:
                return 'BRK'
        elif team == 'CHO' or team == 'CHA' or team == 'CHH':
            if year < 2003:
                return 'CHH'
            elif year > 2014:
                return 'CHO'
            elif 2001 < year < 2005:
                return 'ERROR_CHA_DONT_PLAY_THIS_YEAR'
            else:
                return 'CHA'
        elif team == 'OKC' or team == 'SEA':
            if year < 2009:
                return 'SEA'
            else:
                return 'OKC'


def call_bbr(url, wanted_stats, http_delay: float = 1):
    print(url)

    sleep(http_delay)
    html = urlopen(url)
    soup = BeautifulSoup(html, features="lxml")

    titles = [th.getText() for th in soup.findAll('tr')[1].findAll('th')]

    indexes = []
    for s in wanted_stats:
        indexes.append(titles.index(s))

    data = soup.findAll('tr')[1:]
    rows = [[td.getText() for td in data[i].findAll('td')] for i in range(len(data))]

    result = []
    for stat in wanted_stats:
        stat_list = []
        for row in rows[1:(len(rows))]:
            if len(row) > 0:
                stat_list.append(row[titles.index(stat) - 1])

        result.append(stat_list)

    return result


def get_url(raw_url, team, year):
    return raw_url.replace("TEAM_TO_CHANGE", str(team)).replace("YEAR_TO_CHANGE", str(year))


def main():
    configuration = config.load()

    years = configuration.years
    teams = configuration.teams

    wanted_stats = []
    imported_raw_urls = []
    for request in configuration.requests.values():
        wanted_stats.append(request.stats)
        imported_raw_urls.append(request.url)

    url_lists = []
    for idx, u in enumerate(imported_raw_urls):
        url_lists.append([])
        for y in years:
            for t in range(len(teams)):
                url_lists[idx % len(imported_raw_urls)].append(get_url(u, get_team_acronym(teams[t], y), y))

    # --------------------------------------------------------------------------------

    data = []

    for idx, url_list in enumerate(url_lists):
        for url in url_list:
            print(wanted_stats[idx])
            data.append(call_bbr(url, wanted_stats[idx], configuration.http_delay))

    # --------------------------------- EXCEL EXPORT ---------------------------------

    excel_file = Workbook()

    for idx in range(len(wanted_stats)):
        for year in range(len(years)):
            excel_file.create_sheet(str(years[year]) + '_' + str(idx))
            year_sheet = excel_file[str(years[year]) + '_' + str(idx)]

            for t in range(len(teams)):
                year_sheet.cell(1, len(wanted_stats[idx]) * t + 1, teams[t])

            for sample in range(int(len(data) / len(years))):
                sample += int(len(data) * (year / len(years)))
                for h in range(len(wanted_stats[idx])):
                    year_sheet.cell(2, (h + len(wanted_stats[idx]) * sample) %
                                    int(len(wanted_stats[idx]) * len(data) / len(years)) + 1, wanted_stats[idx][h])

                for column in range(len(data[sample])):
                    for line in range(len(data[sample][column])):
                        year_sheet.cell(line + 3, (column + len(wanted_stats[idx]) * sample) %
                                        int(len(wanted_stats[idx]) * len(data) / len(years)) + 1,
                                        float(data[sample][column][line]))

    excel_file.remove(excel_file["Sheet"])

    file_path = os.path.join(os.path.dirname(__file__), PATH, "DATA" +
                             datetime.now().strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx")
    excel_file.save(file_path)

    print("Saved file: " + file_path)

    # --------------------------------------------------------------------------------


if __name__ == '__main__':
    main()
